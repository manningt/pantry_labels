#!/usr/bin/env python3
'''
2 functions:
1. make_guests_per_caller_lists(in_filename) -> Caller_lists
2. make_caller_pdfs(caller_mapping_dict, guest_dict, date_str)

The input file is an Excel file with 3 sheets: 'guest-to-caller', 'callers', 'guests'
See class definition below for the NamedTuple Caller_lists returned by make_guests_per_caller_lists(in_filename)

The output is a PDF file for each caller with a list of guests for the next Friday.

'''

import sys, os
try:
   from openpyxl import load_workbook
   import csv
   from fpdf import FPDF
except Exception as e:
   print(e)
   sys.exit(1)

import argparse
from pathlib import Path

# from flask import current_app
# from typing import NamedTuple  #not to be confused with namedtuple in collections

# full_guest_dict has item count
def make_full_guest_dict(in_filename):
   guest_dict = {}

   with open(in_filename, newline='') as csvfile:
      reader = csv.DictReader(csvfile)
      # print(f"{reader.fieldnames=}")
      row_count = 0
      for row in reader:
         # full_name = row['Client'].strip()
         try:
            name_key = f"{row['Client'].split(',')[1]}_{row['Client'].split(',')[0]}".strip()
         except:
            print(f"Warning: row {row_count} has no name.")
            continue

         try:
            guest_dict[name_key] = int(float(row['Total Quantity']))
         except ValueError:
            print(f"Warning: {name_key} has no item count.")
            guest_dict[name_key] = 1
         # if row_count == 2:
         #    break
         row_count += 1
   return guest_dict

def make_guest_list(in_filename, guest_dict):
   guest_list = []

   with open(in_filename, newline='') as csvfile:
      reader = csv.DictReader(csvfile)
      # print(f"{reader.fieldnames=}")
      row_count = 0
      for row in reader:
         name_key = row['First'] + '_' + row['Last'].replace('*', '')
         if name_key in guest_dict:
            item_count = guest_dict[name_key]
         else:
            item_count = 0
            print(f"Warning: {name_key} not found in guest_dict.")
         guest_list.append((row['First'], row['Last'], row['Route or Pickup Time'], item_count))
         # print(f"{guest_list[-1]=}")
         # if row_count == 2:
         #    break
         row_count += 1
   return guest_list

def item_count_to_label_count(item_count):
   limits = [0, 9, 17, 25, 32, 40, 49,57,67,73,82,90,100,107,112,139,200]
   if item_count > 200:
      return 16
   for i in range(len(limits)-1):
      if item_count > limits[i] and item_count <= limits[i+1]:
         return i + 1

def make_label_pdfs(guest_list, delivery_type, out_pdf_path):
   # PDF writing examples:
   #  https://medium.com/@mahijain9211/creating-a-python-class-for-generating-pdf-tables-from-a-pandas-dataframe-using-fpdf2-c0eb4b88355c
   #  https://py-pdf.github.io/fpdf2/Tutorial.html
   route_font_size = 28 # allows longer names
   name_font_size = 36
   label_count_font_size = 12
   label_height = 144 #points
   label_width = 288
   number_of_labels = 0

   try:
      pdf = FPDF(orientation="L", unit="pt", format=(label_height,label_width))
      pdf.set_margins(0, 6, 0) #left, top, right in points
      pdf.set_auto_page_break(auto=False)
      pdf.set_font("Helvetica", "B") # Arial not available in fpdf2
      for row in guest_list:
         label_count = item_count_to_label_count(row[3])
         for i in range(label_count):
            pdf.add_page()
            # if row[2] is a time, then don't print it; only print if it's a route
            if delivery_type:
               pdf.set_font_size(route_font_size)
               pdf.cell(0, None, f"{row[2]}", align="L")
               pdf.line(0, 36, label_width, 36) # line from left to right
            pdf.ln(route_font_size+10)
            pdf.set_font_size(name_font_size)
            pdf.cell(0, None, f"{row[0]}", align="C")
            pdf.ln(name_font_size+4)
            pdf.cell(0, None, f"{row[1]}", align="C")
            pdf.ln(name_font_size+4)
            pdf.set_font_size(label_count_font_size)
            pdf.cell(0, None, f"{i+1} of {label_count}", align="R")
            number_of_labels += 1
      pdf.output(out_pdf_path)

   except Exception as e:
      # try:
      #    current_app.logger.warning(f"PDF for {guest} failed: {e}")
      # except:
      print(f"PDF for {guest_list[0]} failed: {e}")

   return number_of_labels


def test_label_pdfs(out_pdf_path):
   route_font_size = 28 # allows longer names
   name_font_size = 36
   item_count_font_size = 12
   try:
      pdf = FPDF(orientation="L", unit="pt", format=(144,288)) # default units are mm; heigth, width are in points - 72 points = 1 inch
      pdf.set_margins(0, 10, 0) #left, top, right in points
      pdf.set_auto_page_break(auto=False)
      pdf.set_font("Helvetica", "B") # Arial not available in fpdf2
      pdf.add_page()
      pdf.set_font_size(route_font_size)
      pdf.cell(0, None, f"Route67890123456789012", align="L", border=1)
      pdf.ln(route_font_size+8)
      pdf.set_font_size(name_font_size)
      pdf.cell(0, None, f"First", align="C", border=1)
      pdf.ln(name_font_size+4)
      pdf.cell(0, None, f"Last", align="C", border=1)
      pdf.ln(name_font_size+4)
      pdf.set_font_size(item_count_font_size)
      pdf.cell(0, None, f"1 of 8", align="R", border=1)
      pdf.output(out_pdf_path)

   except Exception as e:
      print(f"PDF for test_label failed: {e}")


if __name__ == "__main__":

   if 0:
      test_array = [1,11,17,25,32,40,49,57,67,73,82,90,100,107,112,139,200, 201]
      for item_count in test_array:
         print(f"{item_count} is {item_count_to_label_count(item_count)} labels")   
      sys.exit(0)

   if 0:
      test_label_pdfs("/tmp/test_label.pdf")
      sys.exit(0)

   argParser = argparse.ArgumentParser()
   argParser.add_argument("orders_filename", type=str, help="input filename with path")
   argParser.add_argument("friday_pickups_filename", type=str, help="input filename with path")
   argParser.add_argument("saturday_pickups_filename", type=str, help="input filename with path")
   argParser.add_argument("delivery_filename", type=str, help="input filename with path")

   args = argParser.parse_args()

   if args.orders_filename is None:
      sys.exit("Missing orders_filename.")
   if not Path(args.orders_filename).is_file():
      sys.exit("orders_filename is not a file.")
   full_guest_dict = make_full_guest_dict(args.orders_filename)
   if len(full_guest_dict) == 0:
      sys.exit("Failure: orders_filename had no guests.")
   if 0:
      print(f"{full_guest_dict=}")
      sys.exit(0)
   
   guest_filename_list = []
   if args.friday_pickups_filename is None:
      sys.exit("Missing friday_pickups_filename.")
   elif Path(args.friday_pickups_filename).is_file():
      guest_filename_list.append(args.friday_pickups_filename)
   else:
      sys.exit("friday_pickups_filename is not a file.")

   if args.saturday_pickups_filename is None:
      sys.exit("Missing saturday_pickups_filename.")
   elif Path(args.saturday_pickups_filename).is_file():
      guest_filename_list.append(args.saturday_pickups_filename)
   else:
      sys.exit("saturday_pickups_filename is not a file.")

   if args.delivery_filename is None:
      sys.exit("Missing delivery_filename.")
   elif Path(args.delivery_filename).is_file():
      guest_filename_list.append(args.delivery_filename)
   else:
      sys.exit("delivery_filename is not a file.")


   guest_lists = [None] * len(guest_filename_list)
   for i in range(len(guest_filename_list)):
      guest_lists[i] = make_guest_list(guest_filename_list[i], full_guest_dict)
      if len(guest_lists[i]) == 0:
         print(f"Failure: guest_lists {i} had no guests.")
         sys.exit(1)
      # print(f"{guest_lists[i][0][2]}") list - row - tuple index is the route or pickup time
      if (':' in guest_lists[i][0][2] and (' AM' in guest_lists[i][0][2] or ' PM' in guest_lists[i][0][2])):
         delivery_type = False
         type = "pickup"
      else:
         delivery_type = True
         type = "delivery"
      number_of_labels = make_label_pdfs(guest_lists[i], delivery_type, f"/tmp/guest_list_{i}.pdf")
      print(f"guest_list {i} ({type=}) has {len(guest_lists[i])} guests and {number_of_labels} labels.")
